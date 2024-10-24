import logging
from pathlib import Path
from datetime import datetime
import pickle
import sys
import subprocess
import pkg_resources
import time
import joblib
from openpyxl.formatting.rule import ColorScaleRule


from sklearn.base import BaseEstimator, TransformerMixin
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score, confusion_matrix
from sklearn.neighbors import NearestNeighbors
from sklearn.pipeline import Pipeline
import pandas as pd
import xgboost as xgb
import matplotlib.pyplot as plt

import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side
import shap
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import PatternFill, NamedStyle


# Function to install a package if it's not already installed
def install_package(package):
    try:
        pkg_resources.get_distribution(package)
    except pkg_resources.DistributionNotFound:
        subprocess.check_call([sys.executable, "-m", "pip", "install", package])

install_package("openpyxl")
install_package("shap")

import numpy as np
import shap
# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Paths and file names
data_dir = Path(r"C:\Users\16036\OneDrive\Desktop\Investor Sight\Vulnerability\Model")
stata_dir = data_dir
output_dir = data_dir
data_file = "final_dataset3.dta"

# Constants
TARGET_VAR = "F1_factset_diligent_HF"
INDUSTRY_VAR = "sic_1_digit"
INDUSTRY_VAR2 = "sic_2_digit"

CUTOFF_YEAR = 2014
NEIGHBORS_MAPPING = {
    9: 7, # Public Administration
    8: 9, # Services
    2: 17, # Construction
    4: 11, # Transportation
    6: 14, # Retail Trade
    7: 10, # Finance
    3: 7, # Manufacturing
    1: 15, # Mining
    5: 3, # Wholesale Trade
    0: 4 # Agriculture
}

NEIGHBORS_ALL = 1

# SIC 1 Digit mapping
sic_mapping = {
    0: "Agriculture",
    1: "Mining",
    2: "Construction",
    3: "Manufacturing",
    4: "Transportation, Communications, Electric, Gas, and Sanitary Services",
    5: "Wholesale Trade",
    6: "Retail Trade",
    7: "Finance, Insurance, and Real Estate",
    8: "Services",
    9: "Public Administration",
}

sic_mapping2 = {
    1: "Agricultural Production - Crops",
    2: "Agricultural Production - Livestock and Animal Specialties",
    7: "Agricultural Services",
    8: "Forestry",
    9: "Fishing, Hunting and Trapping",
    10: "Metal Mining",
    12: "Coal Mining",
    13: "Oil and Gas Extraction",
    14: "Mining and Quarrying of Nonmetallic Minerals, Except Fuels",
    15: "Construction - General Contractors & Operative Builders",
    16: "Heavy Construction, Except Building Construction, Contractor",
    17: "Construction - Special Trade Contractors",
    20: "Food and Kindred Products",
    21: "Tobacco Products",
    22: "Textile Mill Products",
    23: "Apparel, Finished Products from Fabrics & Similar Materials",
    24: "Lumber and Wood Products, Except Furniture",
    25: "Furniture and Fixtures",
    26: "Paper and Allied Products",
    27: "Printing, Publishing and Allied Industries",
    28: "Chemicals and Allied Products",
    29: "Petroleum Refining and Related Industries",
    30: "Rubber and Miscellaneous Plastic Products",
    31: "Leather and Leather Products",
    32: "Stone, Clay, Glass, and Concrete Products",
    33: "Primary Metal Industries",
    34: "Fabricated Metal Products",
    35: "Industrial and Commercial Machinery and Computer Equipment",
    36: "Electronic & Other Electrical Equipment & Components",
    37: "Transportation Equipment",
    38: "Measuring, Photographic, Medical, & Optical Goods, & Clocks",
    39: "Miscellaneous Manufacturing Industries",
    40: "Railroad Transportation",
    41: "Local & Suburban Transit & Interurban Highway Transportation",
    42: "Motor Freight Transportation",
    43: "United States Postal Service",
    44: "Water Transportation",
    45: "Transportation by Air",
    46: "Pipelines, Except Natural Gas",
    47: "Transportation Services",
    48: "Communications",
    49: "Electric, Gas and Sanitary Services",
    50: "Wholesale Trade - Durable Goods",
    51: "Wholesale Trade - Nondurable Goods",
    52: "Building Materials, Hardware, Garden Supplies & Mobile Homes",
    53: "General Merchandise Stores",
    54: "Food Stores",
    55: "Automotive Dealers and Gasoline Service Stations",
    56: "Apparel and Accessory Stores",
    57: "Home Furniture, Furnishings and Equipment Stores",
    58: "Eating and Drinking Places",
    59: "Miscellaneous Retail",
    60: "Depository Institutions",
    61: "Nondepository Credit Institutions",
    62: "Security & Commodity Brokers, Dealers, Exchanges & Services",
    63: "Insurance Carriers",
    64: "Insurance Agents, Brokers and Service",
    65: "Real Estate",
    67: "Holding and Other Investment Offices",
    70: "Hotels, Rooming Houses, Camps, and Other Lodging Places",
    72: "Personal Services",
    73: "Business Services",
    75: "Automotive Repair, Services and Parking",
    76: "Miscellaneous Repair Services",
    78: "Motion Pictures",
    79: "Amusement and Recreation Services",
    80: "Health Services",
    81: "Legal Services",
    82: "Educational Services",
    83: "Social Services",
    84: "Museums, Art Galleries and Botanical and Zoological Gardens",
    86: "Membership Organizations",
    87: "Engineering, Accounting, Research, and Management Services",
    88: "Private Households",
    89: "Services, Not Elsewhere Classified",
    91: "Executive, Legislative & General Government, Except Finance",
    92: "Justice, Public Order and Safety",
    93: "Public Finance, Taxation and Monetary Policy",
    94: "Administration of Human Resource Programs",
    95: "Administration of Environmental Quality and Housing Programs",
    96: "Administration of Economic Programs",
    97: "National Security and International Affairs",
    99: "Nonclassifiable Establishments"
}



explanation_rules = {
    "accrual": "High",
    "activist_share_pct": "High",
    "activist_support_share_pct": "High",
    "adv_sale": "High",
    "Agency_count": "High",
    "at": "Low",
    "at_turn": "Low",
    "avg_ess": "High",
    "avg_dir_tenure": "High",
    "avg_networksize": "Low",
    "avg_noquals": "Low",
    "avg_sentiment": "n/a", #TO EDIT
    "avg_support_director": "Low",
    "avg_support_management": "Low",
    "avg_support_pay": "Low",
    "avg_totnolstdbrd": "High",
    "avg_totnoothlstdbrd": "Low",
    "avg_totnounlstdbrd": "High",
    "avg_volume": "High",
    "board_independent_ratio": "Low",
    "board_size": "High",
    "boardattendance": "Low",
    "boardbackgroundandskills": "Low",
    "boardmeetingattendanceaverage": "Low",
    "boardmemberlongterm": "Low",
    "busy_dir_pct": "High",
    "capx": "High",
    "cash_conversion": "High",
    "cash_holding": "High",
    "cash_ratio": "High",
    "cboard": "Low",
    "ceo_duality": "High",
    "ceo_female": "High",
    "ceo_overpay": "High",
    "ceo_turnover": "High",
    "ceo_underpay": "Low",
    "ceocompensationlinktotsr": "Low",
    "cfm": "Low",
    "Communal_count": "Low",
    "companycrossshareholding": "Low",
    "CompetitorOrient_count": "High",
    "Confidence_count": "Low",
    "ctype1": "High",
    "ctype18": "High",
    "ctype2": "Low",
    "ctype21": "Low",
    "ctype4": "Low",
    "curr_ratio": "Low",
    "CustomerOrient_count": "Low",
    "de_ratio": "High",
    "debt_at": "High",
    "debt_ebitda": "High",
    "debt_invcap": "High",
    "debt_ratio": "High",
    "ded_pct": "High",
    "dem_inst_pct": "Low",
    "differentvotingrightshare": "Low",
    "directorelectionmajority": "Low",
    "dissatisfied_share_pct": "Low",
    "divyield": "Low",
    "donationstotal": "High",
    "dpr": "Low",
    "e_index": "High",
    "emp_growth": "Low",
    "Environmental_count": "High",
    "environmentpillarscore": "Low",
    "equity_invcap": "High",
    "esgcombinedscore": "Low",
    "evm": "Low",
    "Exploitation_count": "Low",
    "Exploration_count": "Low",
    "fcf_ocf": "Low",
    "female_dir_ratio": "High",
    "FinancialWords_count": "Low",
    "governancepillarscore": "Low",
    "gparachute": "Low",
    "gross_margin": "Low",
    "turnover": "Low",
    "Innovativeness_count": "Low",
    "insider_share_pct": "Low",
    "instblockownpct": "High",
    "instown_hhi": "High",
    "instown_perc": "High",
    "int_totdebt": "High",
    "intan_ratio": "Low",
    "intcov_ratio": "Low",
    "inv_turn": "Low",
    "invt_act": "Low",
    "labylw": "Low",
    "lachtr": "Low",
    "limitedshareholderrightstocallm": "Low",
    "litigationexpenses": "High",
    "litigious_count": "High",
    "liwc_achieve": "Low",
    "liwc_Authentic": "Low",
    "liwc_conflict": "High",
    "liwc_focusfuture": "High",
    "liwc_focuspast": "Low",
    "liwc_focuspresent": "Low",
    "liwc_power": "High",
    "liwc_prosocial": "High",
    "liwc_reward": "Low",
    "liwc_risk": "Low",
    "mb_ratio": "Low",
    "mean_buypct": "Low",
    "mean_holdpct": "Low",
    "mean_rating_overall": "Low",
    "mean_rating_senior_leadership": "Low",
    "mean_rec": "Low",
    "mean_sellpct": "High",
    "mkv": "Low",
    "Narcissism_count": "High",
    "nationalitymix": "High",
    "negative_count": "High",
    "npm": "Low",
    "num_bid": "High",
    "num_buy": "Low",
    "num_diversify_bid": "High",
    "num_negative_pr": "High",
    "num_penalty": "High",
    "num_positive_pr": "Low",
    "num_rec": "Low",
    "num_release": "Low",
    "num_securities_suit": "High",
    "num_sell": "High",
    "num_spinoff": "Low",
    "num_stockholder_suit": "High",
    "num_strike": "High",
    "num_type1": "Low",
    "num_type16": "High",
    "num_type21": "High",
    "num_type22": "Low",
    "num_type23": "High",
    "num_type232": "Low",
    "num_type24": "High",
    "num_type25": "High",
    "num_type26": "High",
    "num_type27": "Low",
    "num_type28": "High",
    "num_type31": "High",
    "num_type32": "Low",
    "num_type43": "High",
    "num_type44": "High",
    "num_type46": "Low",
    "num_type47": "High",
    "num_type73": "High",
    "num_type77": "Low",
    "num_type83": "High",
    "numberofboardmeetings": "Low",
    "numinstblockowners": "High",
    "ocf_lct": "Low",
    "Optimism_count": "Low",
    "OTH_Long_count": "Low",
    "OTH_Short_count": "High",
    "pay_turn": "Low",
    "pcf": "Low",
    "pe_op_basic": "Low",
    "performance_pay": "Low",
    "pol_dir_ratio": "Low",
    "politicalcontributions": "High",
    "positive_count": "Low",
    "ppe_ratio": "High",
    "ppill": "Low",
    "Prevention_count": "High",
    "price_sd": "High",
    "num_recall": "High",
    "Promotion_count": "Low",
    "prstkc": "Low",
    "ps": "Low",
    "qix_pct": "High",
    "quick_ratio": "Low",
    "rd_sale": "High",
    "rect_turn": "High",
    "rep_inst_pct": "High",
    "revt": "Low",
    "roa": "Low",
    "roa_2": "Low",
    "roa_3": "Low",
    "roce": "Low",
    "roe": "Low",
    "roe_2": "Low",
    "roe_3": "Low",
    "salarygap": "High",
    "sale_equity": "Low",
    "sale_invcap": "Low",
    "sale_nwc": "Low",
    "shareholdersapprovalstockcom": "Low",
    "shareholdersvoteonexecutivepay": "Low",
    "short_ratio": "High",
    "socialpillarscore": "Low",
    "staff_sale": "High",
    "stock_ret": "Low",
    "stock_ret_3": "Low",
    "stock_ret_5": "Low",
    "supermajor": "Low",
    "supermajorityvoterequirement": "Low",
    "t_inflow_num_emp": "Low",
    "t_inflow_num_senior_emp": "Low",
    "t_num_employees": "High",
    "t_num_senior_employees": "High",
    "t_outflow_num_emp": "High",
    "t_outflow_num_senior_emp": "High",
    "textblob_sentiment": "Low",
    "tra_pct": "High",
    "tradeunionrepresentation": "Low",
    "uncertainty_count": "High",
    "value_penalty": "High",
    "w_cross_cutting": "High",
    "w_environment": "High",
    "w_governance": "High",
    "w_social": "High",
    "xad": "High",
    "xrd": "High",
    "xsga_ratio": "High",
    "yearfounded": "High"
}


# Function to calculate market cap category from ln_mkv
def market_cap_category(ln_mkv):
    mkv = np.exp(ln_mkv)
    if mkv >= 10**10:
        return "Large Cap (>$10B)"
    elif 2 * 10**9 <= mkv < 10**10:
        return "Mid Cap ($2B-$10B)"
    elif 250 * 10**6 <= mkv < 2 * 10**9:
        return "Small Cap ($250M-$2B)"
    elif  0 <= mkv < 250 * 10**6:
        return "Micro Cap (<$250M)"
    else:
        return "Unknown"
'''
def save_global_shap_plot1(model, features, feature_names, output_path):
    explainer = shap.Explainer(model)
    shap_values = explainer(features)
    shap_values.feature_names = feature_names  # Ensure feature names are included
    max_display = len(features[0])
    shap.summary_plot(shap_values, features=features, feature_names=feature_names, show=False, max_display=max_display)
    plt.savefig(output_path, bbox_inches='tight')
    plt.close()
'''
def save_global_shap_plot1(model, features, feature_names, output_path, selected_features=None):
    explainer = shap.Explainer(model)
    shap_values = explainer(features)

    # Filter the features based on user input if provided
    if selected_features:
        feature_indices = [feature_names.index(feat) for feat in selected_features if feat in feature_names]
        filtered_shap_values = shap_values[:, feature_indices]
        filtered_feature_names = np.array(feature_names)[feature_indices]
    else:
        filtered_shap_values = shap_values
        filtered_feature_names = feature_names

    # Ensure SHAP feature names are set correctly
    filtered_shap_values.feature_names = filtered_feature_names

    # Adjust the number of features to display based on input
    max_display = len(filtered_shap_values[0])

    # Create the summary plot for the selected features
    shap.summary_plot(filtered_shap_values, features=features[:, feature_indices], feature_names=filtered_feature_names, show=False, max_display=max_display)
    plt.savefig(output_path, bbox_inches='tight')
    plt.close()

    print(f"Saved global SHAP plot for selected features to {output_path}")

def save_global_shap_plot1bis(model, features, feature_names, output_path, selected_features=None, selected_labels=None):
    explainer = shap.Explainer(model)
    shap_values = explainer(features)

    # Filter the features based on user input if provided
    if selected_features:
        feature_indices = [feature_names.index(feat) for feat in selected_features if feat in feature_names]
        filtered_shap_values = shap_values[:, feature_indices]
        filtered_feature_names = np.array(feature_names)[feature_indices]
        
        # Use selected_labels if provided and match it with the selected features
        if selected_labels:
            filtered_labels = np.array(selected_labels)[[i for i, feat in enumerate(selected_features) if feat in feature_names]]
        else:
            filtered_labels = filtered_feature_names
    else:
        filtered_shap_values = shap_values
        filtered_feature_names = feature_names
        filtered_labels = feature_names

    # Ensure SHAP feature names are set correctly
    filtered_shap_values.feature_names = filtered_labels

    # Adjust the number of features to display based on input
    max_display = len(filtered_shap_values[0])

    # Create the summary plot for the selected features
    shap.summary_plot(filtered_shap_values, features=features[:, feature_indices], feature_names=filtered_labels, show=False, max_display=max_display)
    plt.savefig(output_path, bbox_inches='tight')
    plt.close()

    print(f"Saved global SHAP plot for selected features with labels to {output_path}")


'''def save_global_shap_plot2(model, features, feature_names, output_path):
    explainer = shap.Explainer(model)
    shap_values = explainer(features)
    shap_values.feature_names = feature_names  # Ensure feature names are included
    max_display = len(features[0])
    shap.plots.bar(shap_values, max_display=max_display)
    plt.savefig(output_path, bbox_inches='tight')
    plt.close()'''

import matplotlib.pyplot as plt
import shap

import matplotlib.pyplot as plt
import shap
import numpy as np

def save_global_shap_plots_in_batches(model, features, feature_names, output_path_prefix, batch_size=20):
    explainer = shap.Explainer(model)
    shap_values = explainer(features)
    shap_values.feature_names = feature_names  # Ensure feature names are included

    num_features = len(feature_names)
    num_batches = (num_features // batch_size) + (1 if num_features % batch_size > 0 else 0)

    for batch_num in range(num_batches):
        start_index = batch_num * batch_size
        end_index = min(start_index + batch_size, num_features)
        
        batch_feature_names = feature_names[start_index:end_index]
        batch_shap_values = shap_values[:, start_index:end_index]
        
        # Ensure SHAP feature names are set correctly for the batch
        batch_shap_values.feature_names = batch_feature_names
        
        # Create the plot for the current batch
        max_display = len(batch_shap_values[0])
        shap.summary_plot(batch_shap_values, features=features[:, start_index:end_index], feature_names=batch_feature_names, show=False, max_display=max_display)

        # Save the plot with a unique name for each batch
        output_path = f"{output_path_prefix}_batch_{batch_num + 1}.png"
        plt.savefig(output_path, bbox_inches='tight')
        plt.close()

        #print(f"Saved SHAP plot for features {start_index + 1} to {end_index} to {output_path}")

        
def save_global_shap_plot2(model, features, feature_names, output_path, top_n=10):
    explainer = shap.Explainer(model)
    shap_values = explainer(features)
    shap_values.feature_names = feature_names  # Ensure feature names are included

    # Get the absolute values of SHAP contributions for ranking
    shap_values_abs_sum = np.abs(shap_values.values).mean(axis=0)

    # Get the indices of the top N most important features based on absolute SHAP values
    top_feature_indices = np.argsort(shap_values_abs_sum)[-top_n:]

    # Select the raw SHAP values for these top features
    top_shap_values = shap_values[:, top_feature_indices]
    top_feature_names = np.array(feature_names)[top_feature_indices]

    # Compute the absolute mean raw SHAP values for each of the top features
    mean_raw_shap_values = np.abs(top_shap_values.values).mean(axis=0)  # Absolute values here

    # Create the plot
    plt.figure(figsize=(10, 6))  # Adjust the size as needed

    # Plot the absolute mean raw SHAP values with the features in horizontal bars
    plt.barh(top_feature_names, mean_raw_shap_values, color="#1f77b4")  # Blue color for the bars

    # Add labels and a title
    plt.xlabel("Mean Absolute SHAP Value")
    plt.ylabel("Feature")
    plt.title(f"Top {top_n} Feature Contributions (Absolute Raw SHAP Values)")

    # Ensure the layout is not cut off
    plt.tight_layout()

    # Save the plot
    plt.savefig(output_path, bbox_inches='tight')
    plt.close()


def save_top_10_shap_interactions(model, features, feature_names, interaction_feature, output_dir):
    explainer = shap.Explainer(model)
    shap_interaction_values = explainer.shap_interaction_values(features)
    
    # Get the index of the interaction_feature (e.g., 'xrd')
    interaction_feature_index = feature_names.index(interaction_feature)
    
    # Extract interaction values for the chosen feature with all others
    interaction_values = shap_interaction_values[:, :, interaction_feature_index]
    
    # Calculate the mean absolute interaction value for each feature with the chosen one
    mean_abs_interactions = np.abs(interaction_values).mean(axis=0)
    
    # Get the indices of the top 10 strongest interactions
    top_10_interactions = np.argsort(mean_abs_interactions)[-10:]
    
    for i, feature_index in enumerate(top_10_interactions):
        feature_name = feature_names[feature_index]
        #print(f"Saving interaction plot between '{interaction_feature}' and '{feature_name}'")

        # Create a SHAP dependence plot showing the interaction between the two features
        shap.dependence_plot((interaction_feature_index, feature_index), shap_interaction_values, features, feature_names=feature_names, show=False)
        
        # Save the interaction plot
        output_path = output_dir / f"{interaction_feature}_interaction_{i+1}_{feature_name}.png"
        plt.savefig(output_path, bbox_inches='tight')
        plt.close()

        #print(f"Saved SHAP interaction plot for {interaction_feature} and {feature_name} to {output_path}")


def save_global_shap_plot2(model, features, feature_names, output_path, top_n=10):
    explainer = shap.Explainer(model)
    shap_values = explainer(features)
    shap_values.feature_names = feature_names  # Ensure feature names are included

    # Get the absolute values of SHAP contributions for ranking
    shap_values_abs_sum = np.abs(shap_values.values).mean(axis=0)

    # Get the indices of the top N most important features based on absolute SHAP values
    top_feature_indices = np.argsort(shap_values_abs_sum)[-top_n:]

    # Select the raw SHAP values for these top features
    top_shap_values = shap_values[:, top_feature_indices]
    top_feature_names = np.array(feature_names)[top_feature_indices]

    # Compute the absolute mean raw SHAP values for each of the top features
    mean_raw_shap_values = np.abs(top_shap_values.values).mean(axis=0)  # Absolute values here

    # Create the plot
    plt.figure(figsize=(10, 6))  # Adjust the size as needed

    # Plot the absolute mean raw SHAP values with the features in horizontal bars
    plt.barh(top_feature_names, mean_raw_shap_values, color="#1f77b4")  # Blue color for the bars

    # Add labels and a title
    plt.xlabel("Mean Absolute SHAP Value")
    plt.ylabel("Feature")
    plt.title(f"Top {top_n} Feature Contributions (Absolute Raw SHAP Values)")

    # Ensure the layout is not cut off
    plt.tight_layout()

    # Save the plot
    plt.savefig(output_path, bbox_inches='tight')
    plt.close()

    #print(f"Saved global SHAP plot with top {top_n} features and absolute SHAP values to {output_path}")

# Function to calculate feature influence
def get_feature_influences(model, feature_names):
    importances = model.named_steps["xgb"].feature_importances_
    return dict(zip(feature_names, importances))

# Function to calculate SHAP values for the features
def calculate_shap_values(model, features):
    explainer = shap.TreeExplainer(model)
    shap_values = explainer.shap_values(features)
    return shap.Explanation(values=shap_values, data=features, feature_names=model.feature_names_in_)

def count_targeted_times(gvkey):
    matched_data = pd.read_csv(output_dir / 'targets.csv')
    matched_data_filtered = matched_data[
        (matched_data['gvkey'] == gvkey) & 
        (matched_data['year'] > 2020.0)  
    ]    
    count_hf = matched_data_filtered['F1_num_factset_diligent_HF'].sum()
    #count_proposal = matched_data_filtered['F1_num_shareholder_proposal'].sum()
    count_total = matched_data_filtered['F1_num_shareholder_proposal'].sum()
    #count_ownership = matched_data_filtered['F1_num_ownership_activism'].sum()
    return count_total, count_hf

def calculate_percentiles_or_fallback(df, feature, sic_2_digit, sic_1_digit):
    min_points = 5
    accumulated_df = df[df['sic_2_digit'] == sic_2_digit]

    # Exclude NaN values from percentile calculation
    accumulated_df = accumulated_df[~accumulated_df[feature].isna()]

    # Check if there are enough data points for this SIC 2-digit group
    if len(accumulated_df) >= min_points:
        percentiles = accumulated_df[feature].quantile([0.2, 0.4, 0.6, 0.8])
        #print(f"Percentiles for SIC2 digit {sic_2_digit}: {percentiles}")
        return {
            '20th_percentile': percentiles.loc[0.2],
            '40th_percentile': percentiles.loc[0.4],
            '60th_percentile': percentiles.loc[0.6],
            '80th_percentile': percentiles.loc[0.8]
        }
    else:
        #print(f"Not enough data for SIC2 digit {sic_2_digit}, falling back to SIC1")
        # Step 2: Try using the SIC 1-digit grouping
        accumulated_df = df[df['sic_1_digit'] == sic_1_digit]
        accumulated_df = accumulated_df[~accumulated_df[feature].isna()]

        if len(accumulated_df) >= min_points:
            percentiles = accumulated_df[feature].quantile([0.2, 0.4, 0.6, 0.8])
            return {
                '20th_percentile': percentiles.loc[0.2],
                '40th_percentile': percentiles.loc[0.4],
                '60th_percentile': percentiles.loc[0.6],
                '80th_percentile': percentiles.loc[0.8]
            }
        else:
            # Step 3: Fallback to overall dataset
            #print(f"Fallback to overall dataset for feature {feature}")
            accumulated_df = df[~df[feature].isna()]
            if len(accumulated_df) >= min_points:
                percentiles = accumulated_df[feature].quantile([0.2, 0.4, 0.6, 0.8])
                return {
                    '20th_percentile': percentiles.loc[0.2],
                    '40th_percentile': percentiles.loc[0.4],
                    '60th_percentile': percentiles.loc[0.6],
                    '80th_percentile': percentiles.loc[0.8]
                }
            else:
                return None  # Handle case where no valid data is found


import json

# Global dictionary to store percentiles for all features and SIC codes
all_percentiles = {}
def calculate_percentiles_for_feature(df, feature):
    """Calculate percentiles for a given feature across SIC 2-digit, SIC 1-digit, and overall."""
    min_points = 5

    # Initialize dictionary to store percentiles for this feature
    feature_percentiles = {}

    # Ensure the feature is numeric and drop non-numeric entries
    df[feature] = pd.to_numeric(df[feature], errors='coerce')

    # Step 1: Calculate by SIC 2-digit
    for sic_2 in df['sic_2_digit'].unique():
        df_sic_2 = df[df['sic_2_digit'] == sic_2].dropna(subset=[feature])
        if len(df_sic_2) >= min_points:
            percentiles = df_sic_2[feature].quantile([0.2, 0.4, 0.6, 0.8]).to_dict()
            feature_percentiles[f"SIC_2_{sic_2}"] = percentiles

    # Step 2: Calculate by SIC 1-digit
    for sic_1 in df['sic_1_digit'].unique():
        df_sic_1 = df[df['sic_1_digit'] == sic_1].dropna(subset=[feature])
        if len(df_sic_1) >= min_points:
            percentiles = df_sic_1[feature].quantile([0.2, 0.4, 0.6, 0.8]).to_dict()
            feature_percentiles[f"SIC_1_{sic_1}"] = percentiles

    # Step 3: Calculate for the entire dataset (overall)
    df_overall = df.dropna(subset=[feature])
    if len(df_overall) >= min_points:
        percentiles = df_overall[feature].quantile([0.2, 0.4, 0.6, 0.8]).to_dict()
        feature_percentiles["Overall"] = percentiles

    # Add the collected percentiles to the global dictionary
    all_percentiles[feature] = feature_percentiles

def save_all_percentiles_to_file():
    """Save all percentiles to a JSON file."""
    output_file = output_dir / "all_percentiles.json"
    with open(output_file, 'w') as f:
        json.dump(all_percentiles, f, indent=4)
    #print(f"Saved all percentiles to {output_file}")

def determine_feature_importance(value, percentiles):
    """ Determine importance based on percentiles or return a meaningful fallback """
    if pd.isna(value):
        return ""  # Handle missing data (NaN)
    
    # Ensure percentiles are calculated correctly
    if percentiles is None:
        return "No percentiles"  # Return "No Data" if there are not enough data points
    
    #print(f"Value: {value}, Percentiles: {percentiles}")  # Debug print

    # Determine the category based on the value and the calculated percentiles
    if value < percentiles['20th_percentile']:
        return "Very Low"
    elif percentiles['20th_percentile'] <= value < percentiles['40th_percentile']:
        return "Low"
    elif percentiles['40th_percentile'] <= value < percentiles['60th_percentile']:
        return "Neutral"
    elif percentiles['60th_percentile'] <= value < percentiles['80th_percentile']:
        return "High"
    else:
        return "Very High"



    
def calculate_vulnerability(confidence_score):
    if confidence_score >= 0.8:
        return "Very High"
    if confidence_score >= 0.5:
        return "High"
    elif confidence_score >= 0.2:
        return "Mid"
    else:
        return "Low"
    
def count_hf_activism_past_years(df, gvkey, year, years=3):
        start_year = year - years
        subset = df[(df['gvkey'] == gvkey) & (df['year'] >= start_year) & (df['year'] < year)]
        return subset['F1_factset_diligent_HF'].sum()


headers = ["Company Information", "Activism Indicators", "Company Fundamentals", "Market Performance", "Governance", "Ownership", "Workforce Dynamics", "E&S", "Events", "Communication", "Media Coverage"]
        
subheaders = [
    # Company Information
    [
        "Ticker", "Company", "Sector", "Sub Sector", 
        "Market Capitalization", "Exchange", 
        "Company Region", "Year Founded"
    ],

    # Activism Indicators
    [
        "Vulnerability", "Vulnerability Score", 
        "# of campaigns (prior 3 years)", 
        "Likelihood of new shareholder proposal", 
        "# of proposals (prior 3 years)"
    ],

    # Company Fundamentals
    [
        "accrual", "adv_sale", "at", "at_turn", "capx", "cash_conversion", 
        "cash_holding", "cash_ratio", "cfm", "curr_ratio", "de_ratio", 
        "debt_at", "debt_ebitda", "debt_invcap", "debt_ratio", 
        "equity_invcap", "fcf_ocf", "gross_margin", "int_totdebt", 
        "intan_ratio", "intcov_ratio", "inv_turn", "invt_act", "mb_ratio", 
        "npm", "ocf_lct", "pay_turn", "ppe_ratio", "quick_ratio", "rd_sale", 
        "rect_turn", "revt", "roa", "roa_2", "roa_3", "roce", "roe", 
        "roe_2", "roe_3", "sale_equity", "sale_invcap", "sale_nwc", 
        "staff_sale", "xad", "xrd", "xsga_ratio", "yearfounded"
    ],

    # Market Performance
    [
        "avg_volume", "divyield", "dpr", "evm", "turnover", "mean_buypct", 
        "mean_holdpct", "mean_rec", "mean_sellpct", "mkv", "num_type232", 
        "num_type26", "num_type27", "num_type46", "num_type47", "pcf", 
        "pe_op_basic", "price_sd", "prstkc", "ps", "short_ratio", "stock_ret", 
        "stock_ret_3", "stock_ret_5"
    ],

    # Governance
    [
        "avg_dir_tenure", "avg_networksize", "avg_noquals", "avg_totnolstdbrd", 
        "avg_totnoothlstdbrd", "avg_totnounlstdbrd", "board_independent_ratio", 
        "board_size", "boardmeetingattendanceaverage", "boardmemberlongterm", 
        "busy_dir_pct", "cboard", "ceo_duality", "ceo_female", "ceo_overpay", 
        "ceo_turnover", "ceo_underpay", "ceocompensationlinktotsr", 
        "companycrossshareholding", "ctype1", "ctype18", "ctype2", "ctype21", 
        "ctype4", "differentvotingrightshare", "directorelectionmajority", 
        "e_index", "female_dir_ratio", "governancepillarscore", "gparachute", 
        "labylw", "lachtr", "limitedshareholderrightstocallm", "nationalitymix", 
        "num_rec", "numberofboardmeetings", "performance_pay", "pol_dir_ratio", 
        "ppill", "shareholdersapprovalstockcom", "shareholdersvoteonexecutivepay", 
        "supermajor", "supermajorityvoterequirement"
    ],

    # Ownership
    [
        "activist_share_pct", "activist_support_share_pct", "avg_support_director", 
        "avg_support_management", "avg_support_pay", "ded_pct", "dem_inst_pct", 
        "dissatisfied_share_pct", "insider_share_pct", "instblockownpct", 
        "instown_hhi", "instown_perc", "num_buy", "num_securities_suit", 
        "num_sell", "num_stockholder_suit", "numinstblockowners", "qix_pct", 
        "rep_inst_pct", "tra_pct"
    ],

        # Workforce Dynamics
    [
        "emp_growth", "mean_rating_overall", "mean_rating_senior_leadership", 
        "num_strike", "num_type44", "salarygap", "t_inflow_num_emp", 
        "t_inflow_num_senior_emp", "t_num_employees", "t_num_senior_employees", 
        "t_outflow_num_emp", "t_outflow_num_senior_emp", "tradeunionrepresentation"
    ],

        # E&S (Environmental & Social)
    [
        "donationstotal", "environmentpillarscore", "esgcombinedscore", 
        "litigationexpenses", "num_penalty", "num_type24", "num_type25", 
        "politicalcontributions", "socialpillarscore", "value_penalty"
    ],

        # Events
    [
        "num_bid", "num_diversify_bid", "num_release", "num_spinoff", 
        "num_type1", "num_type16", "num_type21", "num_type22", "num_type23", 
        "num_type28", "num_type31", "num_type32", "num_type43", "num_type73", 
        "num_type77", "num_type83", "num_recall"
    ],

    # Communication
    [
        "Agency_count", "avg_ess", "avg_sentiment", 
        "boardattendance", "boardbackgroundandskills", "Communal_count", 
        "CompetitorOrient_count", "Confidence_count", "CustomerOrient_count", 
        "Environmental_count", "Exploitation_count", "Exploration_count", 
        "FinancialWords_count", "Innovativeness_count", "litigious_count", 
        "liwc_achieve", "liwc_Authentic", "liwc_conflict", 
        "liwc_focusfuture", "liwc_focuspast", "liwc_focuspresent", 
        "liwc_power", "liwc_prosocial", "liwc_reward", 
        "liwc_risk", "Narcissism_count", "negative_count", "num_negative_pr", 
        "num_positive_pr", "Optimism_count", "OTH_Long_count", "OTH_Short_count", 
        "positive_count", "Prevention_count", "Promotion_count", 
        "textblob_sentiment", "uncertainty_count"
    ],

    # Media Coverage
    [
        "w_cross_cutting", "w_environment", "w_governance", "w_social"
    ]
]
    
    
def main():
    test_dataALL = pd.read_csv(output_dir / "dfw1_merged2.csv")
    test_data = test_dataALL[test_dataALL['russell3000'] == 1]

    flattened_subheaders = [item for sublist in subheaders[2:] for item in sublist]

    # Identify binary, 3-unique, and 4-unique value columns
    binary_columns = [col for col in test_dataALL.columns if test_dataALL[col].nunique() == 2]
    three_unique_columns = [col for col in test_dataALL.columns if test_dataALL[col].nunique() == 3]
    four_unique_columns = [col for col in test_dataALL.columns if test_dataALL[col].nunique() == 4]

    print(f"Binary columns: {binary_columns}")
    print(f"3-unique value columns: {three_unique_columns}")
    print(f"4-unique value columns: {four_unique_columns}")

        # Iterate over all features in the dataset
    for feature in test_dataALL.columns:
        if feature not in ["sic_1_digit", "sic_2_digit", "gvkey", "year"] and \
                feature not in binary_columns and \
                feature not in three_unique_columns and \
                feature not in four_unique_columns:
            calculate_percentiles_for_feature(test_dataALL, feature)

    # Save the collected percentiles to a single JSON file
    save_all_percentiles_to_file()

    #test_dataALL['hf_activism_past_3_years'] = test_dataALL.apply(lambda row: count_hf_activism_past_years(test_dataALL, row['gvkey'], row['year']), axis=1)
    
    percentiles = {}
    feature_names_filename_ALL = output_dir / f"feature_names2_sic_ALL_neighbors_{NEIGHBORS_ALL}.pkl"

    with open(feature_names_filename_ALL, 'rb') as feature_file:
            feature_names_ALL = pickle.load(feature_file)

    missing_features = [feature for feature in feature_names_ALL if feature not in test_dataALL.columns]

    if missing_features:
        print(f"The following features are expected by the model but are missing from the weekly predictors data: {missing_features}")
    else:
        print("All expected features are present in the test data.")
    
    # PROP
    
    feature_names_filename_ALL_prop = output_dir / f"feature_names2_sic_ALL_neighbors_{NEIGHBORS_ALL}prop.pkl"

    with open(feature_names_filename_ALL_prop, 'rb') as feature_file:
            feature_names_ALL_prop = pickle.load(feature_file)

    missing_features_prop = [feature for feature in feature_names_ALL_prop if feature not in test_dataALL.columns]

    if missing_features_prop:
        print(f"The following PROPfeatures are expected by the model but are missing from the weekly predictors data: {missing_features_prop}")
    else: 
        print("All expected PROP features are present in the test data.")
    
    # Take a few rows from the test set for prediction
    test_sample = test_data

    # Predict confidence scores
    confidence_scores = []
    predicted_labels = []
    predicted_probabilities = []

    # Find variables in test_dataALL but not in flattened_subheaders
    unmatched_variables = [col for col in flattened_subheaders if col not in test_dataALL.columns]

    # Print the unmatched variables
    if unmatched_variables:
        print("The following variables are present in the clients menu but not in data:")
        for var in unmatched_variables:
            print(f" - {var}")
    else:
        print("All variables in test_dataALL are accounted for in the flattened_subheaders.")

      # Create missing columns and set them as NaN (blank)
    for variable in unmatched_variables:
        test_dataALL[variable] = np.nan
        test_data[variable] = np.nan

    for _, row in test_sample.iterrows():
        industry = row[INDUSTRY_VAR]
        #model_filename = output_dir / f"model2_sic_{int(industry)}_neighbors_{NEIGHBORS_MAPPING[industry]}.pkl"
        #feature_names_filename = output_dir / f"feature_names2_sic_{int(industry)}_neighbors_{NEIGHBORS_MAPPING[industry]}.pkl"
        model_filename = output_dir / f"model2_sic_ALL_neighbors_{NEIGHBORS_ALL}.pkl"
        feature_names_filename = output_dir / f"feature_names2_sic_ALL_neighbors_{NEIGHBORS_ALL}.pkl"
        
        with open(model_filename, 'rb') as model_file:
            model = pickle.load(model_file)

        with open(feature_names_filename, 'rb') as feature_file:
            feature_names = pickle.load(feature_file)
    
        #PROP

        model_filename_prop = output_dir / f"model2_sic_ALL_neighbors_{NEIGHBORS_ALL}prop.pkl"
        feature_names_filename_prop = output_dir / f"feature_names2_sic_ALL_neighbors_{NEIGHBORS_ALL}prop.pkl"
        
        with open(model_filename_prop, 'rb') as model_file:
            model_prop = pickle.load(model_file)

        with open(feature_names_filename_prop, 'rb') as feature_file:
            feature_names_prop = pickle.load(feature_file)
        
        # Align the features with the saved feature names
        features = row[feature_names].values.reshape(1, -1)
        features_prop = row[feature_names_prop].values.reshape(1, -1)
        confidence_score = model.predict_proba(features)[0][1]  # Probability of class 1
        confidence_score_prop = model_prop.predict_proba(features_prop)[0][1]  # Probability of class 1

        predicted_label = model.predict(features)[0]  # Predicted label
        gvkey=row["gvkey"]
        count_total, count_hf = count_targeted_times(gvkey)

        confidence_score_row = {
            "Ticker": row["ticker"],
            "Company": row["conm"],  # Placeholder
            #"Gvkey": row["gvkey"],
            "Sector": sic_mapping[industry],
            "Sub Sector": sic_mapping2[row["sic_2_digit"]],
            "Market Capitalization": market_cap_category(row["ln_mkv"]),
            "Exchange": row["exchg"],  # Placeholder
            "Company Region": row["state"],  # Placeholder
            "Year Founded": row["yearfounded"],
            "Vulnerability": calculate_vulnerability(confidence_score),
            "Vulnerability Score": confidence_score,  # Multiply by 100 for percentage
            "# of campaigns (prior 3 years)": count_hf,
            "Likelihood of new shareholder proposal": confidence_score_prop,  # Placeholder
            "# of proposals (prior 3 years)": count_total,  # Placeholder
        }

        
        for variable in flattened_subheaders:
            # Skip non-numeric or irrelevant columns
            if variable in ["year", "gvkey", "sic_1_digit", "sic_2_digit"]:
                continue

            # Get the variable's value from the row or assign NaN if not available
            variable_value = row.get(variable, np.nan)

            # Handle binary, 3-unique, and 4-unique columns with NaN check
            if pd.isna(variable_value):
                value = ""  # Assign empty string for NaN values
            elif variable in binary_columns:
                value = "High" if variable_value else "Low"
            elif variable in three_unique_columns:
                levels = ["Low", "Neutral", "High"]
                try:
                    index = int(variable_value)
                    value = levels[index] if 0 <= index < len(levels) else "Unknown"
                except (ValueError, TypeError):
                    value = "Unknown"
            elif variable in four_unique_columns:
                levels = ["Very Low", "Low", "High", "Very High"]                                                                                                                                                             
                try:
                    index = int(variable_value)
                    value = levels[index] if 0 <= index < len(levels) else "Unknown"
                except (ValueError, TypeError):
                    value = "Unknown"
            else:
                # Calculate percentiles for unflagged columns
                percentiles = calculate_percentiles_or_fallback(
                    test_data, variable,
                    sic_2_digit=row['sic_2_digit'],
                    sic_1_digit=row['sic_1_digit']
                )
                # Determine feature importance using the percentiles
                value = determine_feature_importance(variable_value, percentiles)

            # Store the calculated feature importance
            confidence_score_row[variable] = value

        # Append the current row's confidence score and other data
        confidence_scores.append(confidence_score_row)
        predicted_labels.append(predicted_label)
        predicted_probabilities.append(confidence_score)

    # Create DataFrame
    confidence_scores_df = pd.DataFrame(confidence_scores)
    print(confidence_scores_df)
    confidence_scores_df.to_excel(data_dir/'output.xlsx')
    # Save to Excel with formatting
    output_file = output_dir / "model_output5.xlsx"

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        confidence_scores_df.to_excel(writer, sheet_name="output", index=False)

        # Load the workbook and select the active worksheet
        workbook = writer.book
        worksheet = workbook.active

        #print(worksheet)
        worksheet.sheet_view.showGridLines = False

        # Set the font to Aptos Light, size 11
        default_font = Font(name='Aptos Light', size=11)
        for row in worksheet.iter_rows():
            for cell in row:
                cell.font = default_font

        # Merge and format header cells
        header_font = Font(bold=True, name='Aptos Light', size=11)
        subheader_font = Font(name='Aptos Light', size=11)
        alignment = Alignment(horizontal="center", vertical="center")

        # Borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=10, max_col=10):
            for cell in row:
                try:
                    cell.value = round(float(cell.value) * 100, 2)
                except ValueError:
                   continue


        row_to_replace = 1
        start_col = 14

        '''# Step 1: Remove columns in the spreadsheet that are not in flattened_subheaders
        columns_to_remove = []

        for col_idx in range(start_col, worksheet.max_column + 1):
            feature_name = worksheet.cell(row=row_to_replace, column=col_idx).value
            
            # If the feature is not in the flattened subheaders, mark it for removal
            if feature_name not in flattened_subheaders:
                columns_to_remove.append(col_idx)
                #print(f"Marked column {col_idx} (feature: {feature_name}) for removal because it's not in flattened_subheaders.")

        # Remove marked columns from the sheet
        for col_idx in reversed(columns_to_remove):  # Reverse to avoid index shifting issues
            worksheet.delete_cols(col_idx)
            #print(f"Removed column {col_idx} (not in subheaders).")

        #print(f"Deleted features that were not in subheaders: {columns_to_remove}")
        
        # Step 2: Remove subheader features that are not in feature_names
        # Track features from subheaders that are not found in feature_names
        subheaders_to_remove = []

        for subheader in flattened_subheaders:
            if subheader not in feature_names:
                subheaders_to_remove.append(subheader)
                #print(f"Marked subheader '{subheader}' for removal because it's not in feature names.")

        # Remove these subheaders from the original subheader structure while maintaining the overall structure
        for i, sublist in enumerate(subheaders):
            subheaders[i] = [item for item in sublist if item not in subheaders_to_remove]

        # Print the list of removed subheader features
        #print(f"Removed subheader features that were not found in feature names: {subheaders_to_remove}")
        
        #print(f"Headers: {headers}")
        #print(f"Subheaders: {subheaders}")
        '''
        
        pre_formatted_data = pd.DataFrame(worksheet.values)
        pre_formatted_data.to_csv(data_dir / "pre_formatting.csv", index=False)
        print("Data after pre formatting:")
        print(pre_formatted_data.head())

        col_start = 1  # Start from the first column
        for header, subheader in zip(headers, subheaders):
            # Calculate the end column for each header section
            col_end = col_start + len(subheader) - 1

            # Merge cells across the header section
            worksheet.merge_cells(
                start_row=1, start_column=col_start, end_row=1, end_column=col_end
            )
            cell = worksheet.cell(row=1, column=col_start)
            cell.value = header  # Assign the header name
            cell.font = header_font  # Apply header font style
            cell.alignment = alignment  # Center-align
            cell.border = Border(bottom=Side(style='thin'))  # Add bottom border

            # Iterate over the columns for the subheaders
            for i, subheader_name in enumerate(subheader):
                sub_cell = worksheet.cell(row=2, column=col_start + i)
                sub_cell.value = subheader_name  # Assign subheader value
                sub_cell.font = subheader_font  # Apply subheader font
                sub_cell.alignment = alignment  # Center-align subheader
                sub_cell.border = thin_border  # Apply border

            # Move the starting column for the next header section
            col_start = col_end + 1

        formatted_data = pd.DataFrame(worksheet.values)
        formatted_data.to_csv(data_dir / "after_formatting.csv", index=False)
        print("Data after formatting:")
        print(formatted_data.head())

        # Make the "Ticker" column bold
        for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row, min_col=1, max_col=1):
            for cell in row:
                cell.font = Font(bold=True, name='Aptos Light', size=11)

        # Center align specific columns
        for col in ["I", "J", "K", "L", "M"]:
            for cell in worksheet[col]:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Autofill the width of each column
        for col_idx, col in enumerate(worksheet.iter_cols(), 1):  # Iterate over columns, starting with index 1
            max_length = 0
            column_letter = get_column_letter(col_idx)  # Convert column index to column letter
    
            for cell in col:
                try:
                    if not isinstance(cell, MergedCell) and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Increase the height of row 2
        worksheet.row_dimensions[2].height = 30

        # Add bottom border to the last row
        last_row = worksheet.max_row
        for cell in worksheet[last_row]:
            cell.border = Border(bottom=Side(style='thin'))

        color_scale_rule = ColorScaleRule(
            start_type='min', start_color='63BE7B',  # Start with green
            mid_type='percentile', mid_value=50, mid_color='FFEB84',  # Midpoint at 65th percentile
            end_type='max', end_color='F8696B'  # End with red
        )

        worksheet.conditional_formatting.add('J3:J{}'.format(last_row), color_scale_rule)

        fill_color = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        # Create a NamedStyle for percentage formatting without decimals
        percent_style = NamedStyle(name="percent_style", number_format="0%", fill=fill_color)

        # Apply the style from G3 to the bottom of the column
        last_row = worksheet.max_row  # Get the last row of the worksheet

        for row in range(3, last_row + 1):
            cell = worksheet[f'L{row}']
            cell.style = percent_style
    
        # Apply feature colors based on the feature_colors dictionary
        min_row = 3
        start_col = 14  # Column N is the 14th column

        print(f"Setting grey color for rows")

        for row_idx in range(min_row, worksheet.max_row + 1):
            # Iterate over feature columns starting from the start_col to the last column
            for col_idx in range(start_col, worksheet.max_column + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.font = Font(color="808080")  # Apply grey color to the font


        min_row = 3
        start_col = 14  # Column N is the 14th column

        for col_idx in range(start_col, worksheet.max_column + 1):
            
            # Look up the feature name from row 2 (header row)
            feature_name = worksheet.cell(row=2, column=col_idx).value

            if feature_name in explanation_rules:
                explanation_rule = explanation_rules[feature_name]

                # Iterate through each row for this column
                for row_idx in range(min_row, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    value = cell.value

                    # Apply colors based on the explanation rule and value
                    if explanation_rule == "High":
                        if value == "Very High":
                            cell.font = Font(color="ff0000")  # Full Red666666 for Very High
                            #print(f"Applied Full Red to {column_name}{row_idx} for Very High value")
                        elif value == "High":
                            cell.font = Font(color="ff7765")  # Light Red for High
                            #print(f"Applied Light Red to {column_name}{row_idx} for High value")
                        elif value == "Low":
                            cell.font = Font(color="09be00")  # Light Green for Low
                            #print(f"Applied Light Green to {column_name}{row_idx} for Low value")
                        elif value == "Very Low":
                            cell.font = Font(color="048000")  # Full Green for Very Low
                            #print(f"Applied Full Green to {column_name}{row_idx} for Very Low value")

                    elif explanation_rule == "Low":
                        if value == "Very Low":
                            cell.font = Font(color="ff0000")  # Full Red for Very Low
                            #print(f"Applied Full Red to {column_name}{row_idx} for Very Low value")
                        elif value == "Low":
                            cell.font = Font(color="ff7765")  # Light Red for Low
                            #print(f"Applied Light Red to {column_name}{row_idx} for Low value")
                        elif value == "High":
                            cell.font = Font(color="09be00")  # Light Green for High
                            #print(f"Applied Light Green to {column_name}{row_idx} for High value")
                        elif value == "Very High":
                            cell.font = Font(color="048000")  # Full Green for Very High
                           #print(f"Applied Full Green to {column_name}{row_idx} for Very High value")
                        
            #else:
                #print(f"No explanation rule found for {feature_name}")

        # Generating the variable_label_mapping dictionary based on the data provided
        variable_label_mapping = {
            "accrual": "Accruals/Average assets",
            "activist_share_pct": "Activist investor ownership",
            "activist_support_share_pct": "Ownership by discontented shareholders",
            "adv_sale": "Advertising expenses/Sales",
            "Agency_count": "Agentic values",
            "at": "Total assets",
            "at_turn": "Asset turnover",
            "avg_ess": "Press release sentiment",
            "avg_dir_tenure": "Average director tenure",
            "avg_networksize": "Director connectedness",
            "avg_noquals": "Director education",
            "avg_risk": "Annual average disclosed overall risk",
            "avg_sentiment": "Annual average disclosed overall sentiment",
            "avg_support_director": "Shareholder support for directors",
            "avg_support_management": "Support for management proposals",
            "avg_support_pay": "Support for executive pay",
            "avg_totnolstdbrd": "Director duties on other listed boards",
            "avg_totnoothlstdbrd": "Director duties on non-profit boards",
            "avg_totnounlstdbrd": "Director duties on non-listed boards",
            "avg_volume": "Stock trading volume",
            "board_independent_ratio": "Board independence",
            "board_size": "Board size",
            "boardattendance": "Publishes board attendance",
            "boardbackgroundandskills": "Publishes director experience",
            "boardmeetingattendanceaverage": "Board attendance",
            "boardmemberlongterm": "Director long-term compensation",
            "busy_dir_pct": "Over-boarded directors",
            "capx": "CAPX",
            "cash_conversion": "Cash conversion cycle",
            "cash_holding": "Cash holding ratio",
            "cash_ratio": "Cash ratio",
            "cboard": "Classified board",
            "ceo_duality": "CEO is board chair",
            "ceo_female": "Female CEO",
            "ceo_overpay": "Unjustified CEO pay increase",
            "ceo_turnover": "CEO turnover",
            "ceo_underpay": "Surprise CEO pay decrease",
            "ceocompensationlinktotsr": "CEO compensation dependent on TSR",
            "cfm": "Cash flow margin",
            "Communal_count": "Communal values",
            "companycrossshareholding": "Anti-takeover cross shareholdings",
            "CompetitorOrient_count": "Competitor focus",
            "Confidence_count": "Managerial confidence",
            "ctype1": "CEO salary",
            "ctype18": "CEO total compensation",
            "ctype2": "CEO bonus",
            "ctype21": "CEO option awards",
            "ctype4": "CEO restricted stock",
            "curr_ratio": "Current ratio",
            "CustomerOrient_count": "Customer focus",
            "de_ratio": "Total debt/Equity",
            "debt_at": "Total debt/Total assets",
            "debt_ebitda": "Total debt/EBITDA",
            "debt_invcap": "Long-term debt/Invested capital",
            "debt_ratio": "Debt ratio",
            "ded_pct": "Dedicated institutional ownership",
            "dem_inst_pct": "Democratic institutional ownership",
            "differentvotingrightshare": "Dual class share structure",
            "directorelectionmajority": "Director majority vote",
            "dissatisfied_share_pct": "Ownership by management supportive shareholders",
            "divyield": "Dividend yield",
            "donationstotal": "Charitable donations",
            "dpr": "Dividend payout ratio",
            "e_index": "Entrenchment index",
            "emp_growth": "Employee growth rate",
            "Environmental_count": "Environmental focus",
            "environmentpillarscore": "Environmental score",
            "equity_invcap": "Common equity/Invested capital",
            "esgcombinedscore": "ESG rating",
            "evm": "Enterprise value multiple",
            "Exploitation_count": "Exploitation orientation",
            "Exploration_count": "Exploration orientation",
            "fcf_ocf": "Free cash flow/Operating cash flow",
            "female_dir_ratio": "Female directors",
            "FinancialWords_count": "Financial focus",
            "governancepillarscore": "Governance score",
            "gparachute": "Golden parachutes",
            "gross_margin": "Gross margin",
            "turnover": "Stock Turnover",
            "Innovativeness_count": "Innovation focus",
            "insider_share_pct": "Insider ownership by executives and board",
            "instblockownpct": "Ownership by institutional blockholders",
            "instown_hhi": "Ownership concentration",
            "instown_perc": "Institutional ownership",
            "int_totdebt": "Interest/Average total debt",
            "intan_ratio": "Intangible assets/Total assets",
            "intcov_ratio": "Interest coverage ratio",
            "inv_turn": "Inventory turnover",
            "invt_act": "Inventory/Current assets",
            "labylw": "Amend bylaws constraints",
            "lachtr": "Amend charter constraints",
            "limitedshareholderrightstocallm": "Special meeting constraints",
            "litigationexpenses": "Litigation expenses incurred",
            "litigious_count": "Legal focus",
            "liwc_achieve": "Achievement motive",
            "liwc_Authentic": "Authenticity",
            "liwc_conflict": "Managerial disagreement",
            "liwc_focusfuture": "Past focus",
            "liwc_focuspast": "Future focus",
            "liwc_focuspresent": "Present focus",
            "liwc_power": "Power motive",
            "liwc_prosocial": "Prosocial focus",
            "liwc_reward": "Reward motive",
            "liwc_risk": "Risk-taking disclosure",
            "mb_ratio": "Market-to-book ratio",
            "mean_buypct": "Average quarterly buy recommendation",
            "mean_holdpct": "Average quarterly hold recommendation",
            "mean_rating_overall": "Employee satisfaction",
            "mean_rating_senior_leadership": "Employee approval of senior leadership",
            "mean_rec": "Average quarterly recommendation",
            "mean_sellpct": "Average annual sell recommendation",
            "mkv": "Market capitalization",
            "Narcissism_count": "Narcissistic displays",
            "nationalitymix": "Foreign directors",
            "negative_count": "Negative business sentiment",
            "npm": "Net profit margin",
            "num_bid": "Acquisitions",
            "num_buy": "Insider stock purchases",
            "num_diversify_bid": "Diversifying acquisitions",
            "num_negative_pr": "Negative press releases",
            "num_penalty": "Number of regulatory violations",
            "num_positive_pr": "Positive press releases",
            "num_rec": "Analyst coverage",
            "num_release": "Product releases",
            "num_securities_suit": "Securities-related lawsuits",
            "num_sell": "Insider stock sell-offs",
            "num_spinoff": "Spin off(s)",
            "num_stockholder_suit": "General shareholder lawsuits",
            "num_strike": "Number of worker strikes",
            "num_type1": "Divestment",
            "num_type16": "Executive/board changes",
            "num_type21": "Downsizing",
            "num_type22": "Strategic alliances",
            "num_type23": "Client announcements",
            "num_type232": "Buyback transaction announcements",
            "num_type24": "Regulatory agency inquiries",
            "num_type25": "Lawsuits and legal issues",
            "num_type26": "Company lowered corporate guidance",
            "num_type27": "Company raised corporate guidance",
            "num_type28": "Earnings announcements",
            "num_type31": "Business expansions",
            "num_type32": "Business reorganizations",
            "num_type43": "Restatements of operating results",
            "num_type44": "Labor-related announcements",
            "num_type46": "Dividend increases",
            "num_type47": "Dividend decreases",
            "num_type73": "Impairments or write-offs",
            "num_type77": "Changes in company bylaws",
            "num_type83": "Private placements",
            "numberofboardmeetings": "Number of board meetings",
            "numinstblockowners": "Number of blockholders",
            "ocf_lct": "Operating cash flow/Current liabilities",
            "Optimism_count": "Managerial optimism",
            "OTH_Long_count": "Long-term horizon",
            "OTH_Short_count": "Short-term horizon",
            "pay_turn": "Payables turnover",
            "pcf": "Stock price/Cash flow from operations",
            "pe_op_basic": "Stock price/Operating earnings",
            "performance_pay": "CEO performance based pay ratio",
            "pol_dir_ratio": "Political directors",
            "politicalcontributions": "Political donations",
            "positive_count": "Positive business sentiment",
            "ppe_ratio": "PPE ratio",
            "ppill": "Poison pill",
            "Prevention_count": "Prevention focus",
            "price_sd": "Stock return volatility",
            "num_recall": "Mass product recall or withdrawal",
            "Promotion_count": "Promotion focus",
            "prstkc": "Purchase of common and preferred stocks",
            "ps": "Stock price/Sales",
            "qix_pct": "Quasi-index institutional ownership",
            "quick_ratio": "Quick ratio (Acid test)",
            "rd_sale": "R&D expenditures/Sales",
            "rect_turn": "Receivables turnover",
            "rep_inst_pct": "Republican institutional ownership",
            "revt": "Total revenue",
            "roa": "ROA",
            "roa_2": "ROA change previous two quarters",
            "roa_3": "ROA change over previous three quarters",
            "roce": "Return on capital employed",
            "roe": "ROE",
            "roe_2": "ROE change previous two quarters",
            "roe_3": "ROE change previous three quarters",
            "russell1000": "Whether a firm is part of Russell 1000 index",
            "russell2000": "Whether a firm is part of Russell 2000 index",
            "russell3000": "Whether a firm is part of Russell 3000 index",
            "salarygap": "CEO-workforce salary gap",
            "sale_equity": "Sales/Stockholders equity",
            "sale_invcap": "Sales/Invested capital",
            "sale_nwc": "Sales/Working capital",
            "shareholdersapprovalstockcom": "Shareholder approval for executive compensation",
            "shareholdersvoteonexecutivepay": "Shareholder vote on executive compensation",
            "short_ratio": "Short ratio",
            "socialpillarscore": "Social score",
            "staff_sale": "Labor expenses/Sales",
            "state": "Headquarters (state)",
            "stock_ret": "Quarterly stock returns",
            "stock_ret_3": "Compound average stock returns in prior three quarters",
            "stock_ret_5": "Compound average stock returns in prior five quarters",
            "supermajor": "Supermajority to approve merger",
            "supermajorityvoterequirement": "Supermajority vote requirement",
            "t_inflow_num_emp": "Inflow of employees",
            "t_inflow_num_senior_emp": "Inflow of senior employees",
            "t_num_employees": "Number of employees",
            "t_num_senior_employees": "Number of senior employees",
            "t_outflow_num_emp": "Outflow of employees",
            "t_outflow_num_senior_emp": "Outflow of senior employees",
            "textblob_sentiment": "Analyst sentiment",
            "tra_pct": "Transient institutional ownership",
            "tradeunionrepresentation": "Trade union representation",
            "uncertainty_count": "Managerial uncertainty",
            "value_penalty": "Cost of regulatory violations",
            "w_cross_cutting": "Cross-cutting ESG issue coverage",
            "w_environment": "Environment issue coverage",
            "w_governance": "Governance issue coverage",
            "w_social": "Social issue coverage",
            "xad": "Advertising expenditures",
            "xrd": "R&D expenditures",
            "xsga_ratio": "SGA/Revenue",
            "yearfounded": "Year established"
        }

        pre_final_data = pd.DataFrame(worksheet.values)
        pre_final_data.to_csv(data_dir / "pre_final.csv", index=False)
        print("Data after pre final:")
        print(pre_final_data.head())


        # Assuming variable_label_mapping is the dictionary created earlier
        start_col = 14  # Column N is the 14th column
        row_to_replace = 2  # Row where the variable names are currently located

        for col_idx in range(start_col, worksheet.max_column + 1):
            # Get the current variable name from row 3
            variable_name = worksheet.cell(row=row_to_replace, column=col_idx).value
            
            # Look for the variable name in the mapping and replace if found
            if variable_name in variable_label_mapping:
                worksheet.cell(row=row_to_replace, column=col_idx).value = variable_label_mapping[variable_name]
                print(f"Replaced variable name '{variable_name}' with label '{variable_label_mapping[variable_name]}' in column {col_idx}")
            else:
                print(f"No mapping found for variable name '{variable_name}' in column {col_idx}")

        post_final_data = pd.DataFrame(worksheet.values)
        post_final_data.to_csv(data_dir / "post_final.csv", index=False)
        print("Data after post final:")
        print(post_final_data.head())

    all_features = test_sample[feature_names].values
    global_shap_plot_path1 = output_dir / "global_shap_plot1.png"
    global_shap_plot_path2 = output_dir / "global_shap_plot2.png"

    '''save_global_shap_plot1(model.named_steps['xgb'], all_features, feature_names, global_shap_plot_path1)
    save_global_shap_plot2(model.named_steps['xgb'], all_features, feature_names, global_shap_plot_path2)'''

    save_global_shap_plot2(model.named_steps['xgb'], all_features, feature_names, global_shap_plot_path2, top_n=10)
    selected_features = ['hf_activism_past_3_years', 'activist_share_pct', 'avg_networksize', 'ind_adj_stock_ret', 'ocf_lct', 'short_ratio', 'board_independent_ratio', 't_outflow_num_senior_emp']  
    selected_labels = [variable_label_mapping.get(feature, feature) for feature in selected_features]

    save_global_shap_plot1bis(
        model.named_steps['xgb'], 
        all_features, 
        feature_names, 
        global_shap_plot_path1, 
        selected_features=selected_features, 
        selected_labels=selected_labels
    ) 
    #save_global_shap_plots_in_batches(model.named_steps['xgb'], all_features, feature_names, global_shap_plot_path1)

    #save_top_10_shap_interactions(model.named_steps['xgb'], all_features, feature_names, 'stock_ret', output_dir)

    print(f"Formatted confidence scores saved to {output_file}")
    print(f"Global SHAP plots saved to {global_shap_plot_path1} and {global_shap_plot_path2}")

if __name__ == "__main__":
    main()
